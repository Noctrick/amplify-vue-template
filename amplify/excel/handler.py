import os
import re
import copy
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill, Font
import boto3

# Initialize the S3 client
s3_client = boto3.client('s3')

# Use Lambda's temporary directory for processing
OUTPUT_FOLDER = '/tmp/output'
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def sanitize_filename(name):
    return re.sub(r'[\\/*?:"<>|]', '-', str(name)).strip()

def process_excel(input_file, output_folder):
    wb = openpyxl.load_workbook(input_file)
    # Adjust if your sheet name differs
    sheet = wb["Energiesnoeier"]

    # Define which columns to copy (example: columns A-T)
    columns_to_keep = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"]
    column_indices = [column_index_from_string(col) for col in columns_to_keep]

    # Define header styling
    header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Group rows by value in column D (index 3)
    grouped_rows = {}
    for row in sheet.iter_rows(min_row=2, values_only=False):
        klant_cell = row[3]
        klant_value = klant_cell.value
        if not klant_value:
            continue
        klant_key = sanitize_filename(klant_value)
        grouped_rows.setdefault(klant_key, []).append(row)

    # Process each group and create new Excel file per group
    for klant, rows in grouped_rows.items():
        klant_folder = os.path.join(output_folder, klant)
        os.makedirs(klant_folder, exist_ok=True)

        new_wb = Workbook()
        new_sheet = new_wb.active
        new_sheet.title = "Data"

        max_lengths = {i: 0 for i in range(1, len(column_indices) + 1)}

        # Copy header row (assumed row 1 in the source)
        for idx, col_idx in enumerate(column_indices, start=1):
            src_cell = sheet.cell(row=1, column=col_idx)
            dst_cell = new_sheet.cell(row=1, column=idx, value=src_cell.value)
            dst_cell.fill = header_fill
            dst_cell.font = header_font
            dst_cell.border = copy.copy(src_cell.border)
            dst_cell.alignment = copy.copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format

            header_text = str(src_cell.value) if src_cell.value else ""
            max_lengths[idx] = max(max_lengths[idx], len(header_text))

        # Copy data rows for the current group
        for new_row_idx, src_row in enumerate(rows, start=2):
            for idx, col_idx in enumerate(column_indices, start=1):
                src_cell = src_row[col_idx - 1]
                dst_cell = new_sheet.cell(row=new_row_idx, column=idx, value=src_cell.value)
                dst_cell.font = copy.copy(src_cell.font)
                dst_cell.fill = copy.copy(src_cell.fill)
                dst_cell.border = copy.copy(src_cell.border)
                dst_cell.alignment = copy.copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format

                cell_text = str(src_cell.value) if src_cell.value is not None else ""
                max_lengths[idx] = max(max_lengths[idx], len(cell_text))

        # Auto-adjust column widths
        for idx in range(1, len(column_indices) + 1):
            col_letter = get_column_letter(idx)
            new_sheet.column_dimensions[col_letter].width = max_lengths[idx] + 2

        output_filename = f"Startstanden_{klant}.xlsx"
        output_path = os.path.join(klant_folder, output_filename)
        new_wb.save(output_path)
        print(f"Created file for {klant}: {output_path}")

    return "Processing complete. Files saved."

def lambda_handler(event, context):
    """
    Expected event format:
    {
      "bucket": "your-s3-bucket-name",
      "key": "path/to/input.xlsx"
    }
    """
    bucket = event.get('bucket')
    key = event.get('key')
    if not bucket or not key:
        return {
            'statusCode': 400,
            'body': 'Missing bucket or key in the event'
        }
    
    download_path = '/tmp/input.xlsx'
    s3_client.download_file(bucket, key, download_path)

    # Process the Excel file
    message = process_excel(download_path, OUTPUT_FOLDER)

    # Upload processed files back to S3 under a "processed/" prefix
    for root, dirs, files in os.walk(OUTPUT_FOLDER):
        for filename in files:
            local_path = os.path.join(root, filename)
            s3_key = f"processed/{filename}"
            s3_client.upload_file(local_path, bucket, s3_key)

    return {
        'statusCode': 200,
        'body': message
    }
